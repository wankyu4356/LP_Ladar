"""Excel report generation."""
from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lp_radar.models import Announcement, ScrapeError

YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="0563C1", underline="single")


def generate_excel(
    announcements: list[Announcement],
    errors: list[ScrapeError],
    output_path: Path,
) -> Path:
    """Generate an Excel report with announcements and error sheets."""
    wb = Workbook()

    # --- Announcements Sheet ---
    ws = wb.active
    ws.title = "공고 목록"

    headers = ["기관명", "약칭", "날짜", "제목", "URL", "펀드관련"]
    col_widths = [20, 10, 12, 60, 50, 10]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, ann in enumerate(announcements, 2):
        ws.cell(row=row_idx, column=1, value=ann.source)
        ws.cell(row=row_idx, column=2, value=ann.source_short)
        ws.cell(row=row_idx, column=3, value=ann.date.isoformat() if ann.date else "")
        ws.cell(row=row_idx, column=4, value=ann.title)

        url_cell = ws.cell(row=row_idx, column=5, value=ann.url)
        url_cell.font = LINK_FONT
        url_cell.hyperlink = ann.url

        ws.cell(row=row_idx, column=6, value="Y" if ann.is_fund_related else "N")

        # Highlight fund-related rows
        if ann.is_fund_related:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL

    # Auto-filter
    if announcements:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(announcements) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Summary Sheet ---
    ws_summary = wb.create_sheet("요약")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15

    summary_headers = ["기관명", "전체 공고 수", "펀드관련 공고 수"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Count per source
    source_counts: dict[str, tuple[int, int]] = {}
    for ann in announcements:
        total, fund = source_counts.get(ann.source, (0, 0))
        total += 1
        if ann.is_fund_related:
            fund += 1
        source_counts[ann.source] = (total, fund)

    for row_idx, (source, (total, fund)) in enumerate(sorted(source_counts.items()), 2):
        ws_summary.cell(row=row_idx, column=1, value=source)
        ws_summary.cell(row=row_idx, column=2, value=total)
        ws_summary.cell(row=row_idx, column=3, value=fund)
        if fund > 0:
            for col_idx in range(1, 4):
                ws_summary.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL

    ws_summary.freeze_panes = "A2"

    # --- Errors Sheet ---
    if errors:
        ws_err = wb.create_sheet("에러")
        err_headers = ["기관명", "약칭", "URL", "에러 메시지", "스크린샷"]
        err_widths = [20, 10, 50, 50, 40]

        for col_idx, (header, width) in enumerate(zip(err_headers, err_widths), 1):
            cell = ws_err.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws_err.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, err in enumerate(errors, 2):
            ws_err.cell(row=row_idx, column=1, value=err.source)
            ws_err.cell(row=row_idx, column=2, value=err.source_short)
            ws_err.cell(row=row_idx, column=3, value=err.url)
            ws_err.cell(row=row_idx, column=4, value=err.error_message)
            ws_err.cell(row=row_idx, column=5, value=err.screenshot_path or "")
            for col_idx in range(1, len(err_headers) + 1):
                ws_err.cell(row=row_idx, column=col_idx).fill = RED_FILL

        ws_err.freeze_panes = "A2"

    # --- Metadata ---
    ws_meta = wb.create_sheet("실행 정보")
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 40
    ws_meta.cell(row=1, column=1, value="실행 시간").font = Font(bold=True)
    ws_meta.cell(row=1, column=2, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws_meta.cell(row=2, column=1, value="총 공고 수").font = Font(bold=True)
    ws_meta.cell(row=2, column=2, value=len(announcements))
    ws_meta.cell(row=3, column=1, value="펀드관련 공고 수").font = Font(bold=True)
    ws_meta.cell(row=3, column=2, value=sum(1 for a in announcements if a.is_fund_related))
    ws_meta.cell(row=4, column=1, value="에러 사이트 수").font = Font(bold=True)
    ws_meta.cell(row=4, column=2, value=len(errors))

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
